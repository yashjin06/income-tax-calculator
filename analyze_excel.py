import openpyxl
from openpyxl.utils import get_column_letter

file_path = r'f:\Antigravity\Income Tax Calculator\File for Capital Gain.xlsx'
try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
except Exception as e:
    print(f"Failed to load: {e}")
    exit(1)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"--- Sheet: {sheet_name} ---")
    rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))
    for row in rows:
        print(row)
