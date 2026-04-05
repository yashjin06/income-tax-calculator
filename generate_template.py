import openpyxl
import os

source_file = r'f:\Antigravity\Income Tax Calculator\File for Capital Gain.xlsx'
target_file = r'f:\Antigravity\Income Tax Calculator\public\TaxNova Capital Gains Template.xlsx'
root_temp = r'f:\Antigravity\Income Tax Calculator\TaxNova Capital Gains Template.xlsx'

print(f"Loading {source_file}")
wb = openpyxl.load_workbook(source_file, data_only=False)

def apply_shares_logic(ws):
    print(f"Applying logic and clearing data for sheet: {ws.title}")
    # Formulas and clear data start at row 2, goes to 100
    for i in range(2, 101):
        # Clear user input columns A through G
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{i}'].value = None
            
        # Column H = Net Gain
        ws[f'H{i}'] = f'=IF(AND(D{i}<>"",F{i}<>""),F{i}-D{i}-G{i},"")'
        # Column I = Holding Period
        ws[f'I{i}'] = f'=IF(AND(C{i}<>"", E{i}<>""), (E{i}-C{i}) & " days", "")'
        # Column J = Type of Capital Gain
        ws[f'J{i}'] = f'=IF(OR(C{i}="", E{i}="", A{i}=""), "", IF(LEFT(A{i}, 6)="Listed", IF((E{i}-C{i})>365, "LTCG", "STCG"), IF((E{i}-C{i})>730, "LTCG", "STCG")))'
        # Column K = Rate of Tax
        ws[f'K{i}'] = f'=IF(J{i}="", "", IF(J{i}="LTCG", "12.5%", IF(LEFT(A{i},6)="Listed", "20.0%", "Slab Rate")))'
        # Column L = Tax Value
        ws[f'L{i}'] = f'=IF(OR(H{i}="", K{i}=""), "", IF(K{i}="Slab Rate", "As per Slab", IF(ISNUMBER(VALUE(LEFT(K{i},LEN(K{i})-1))), H{i}*VALUE(LEFT(K{i},LEN(K{i})-1))/100, "")))'

def apply_mf_logic(ws):
    print(f"Applying logic and clearing data for sheet: {ws.title}")
    # Formulas and clear data start at row 2, goes to 100
    for i in range(2, 101):
        # Clear user input columns A through G
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{i}'].value = None
            
        # Column H = Net Gain
        ws[f'H{i}'] = f'=IF(AND(D{i}<>"",F{i}<>""),F{i}-D{i}-G{i},"")'
        # Column I = Holding Period
        ws[f'I{i}'] = f'=IF(AND(C{i}<>"", E{i}<>""), (E{i}-C{i}) & " days", "")'
        # Column J = Type of Capital Gain (For Debt MF, usually STCG from Apr 2023)
        ws[f'J{i}'] = f'=IF(OR(C{i}="", E{i}="", A{i}=""), "", IF(LEFT(A{i},9)="MF (Equit", IF((E{i}-C{i})>365, "LTCG", "STCG"), IF(C{i}>DATE(2023,4,1), "STCG (Fixed)", "STCG")))'
        # Column K = Rate of Tax
        ws[f'K{i}'] = f'=IF(J{i}="", "", IF(J{i}="LTCG", "12.5%", IF(LEFT(A{i},9)="MF (Equit", "20.0%", "Slab Rate")))'
        # Column L = Tax Value
        ws[f'L{i}'] = f'=IF(OR(H{i}="", K{i}=""), "", IF(K{i}="Slab Rate", "As per Slab", IF(ISNUMBER(VALUE(LEFT(K{i},LEN(K{i})-1))), H{i}*VALUE(LEFT(K{i},LEN(K{i})-1))/100, "")))'

if 'Sale of Shares' in wb.sheetnames:
    apply_shares_logic(wb['Sale of Shares'])

if 'Sale of Mutual Funds' in wb.sheetnames:
    apply_mf_logic(wb['Sale of Mutual Funds'])

# Just in case they are named differently
for sheet_name in wb.sheetnames:
    if 'share' in sheet_name.lower() and sheet_name != 'Sale of Shares':
        apply_shares_logic(wb[sheet_name])
    if 'fund' in sheet_name.lower() and 'mutual' in sheet_name.lower() and sheet_name != 'Sale of Mutual Funds':
        apply_mf_logic(wb[sheet_name])

print(f"Saving to {target_file}")
wb.save(target_file)
wb.save(root_temp)
print("Done. Generated new template retaining all original styles, but with sample data cleared.")
